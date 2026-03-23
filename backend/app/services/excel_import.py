"""Parse existing Geraeteuebersicht.xlsx and import into SQLite."""

from __future__ import annotations

import io
import logging
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Expected header columns (matching create_device_list.py)
EXPECTED_HEADERS = [
    "Nr", "Typ", "Bezeichnung", "Modell", "Hersteller", "Standort",
    "Seriennummer", "MAC-Adresse", "IP-Adresse", "Firmware",
    "Integration (HA)", "Netzwerk", "Stromversorgung", "AIN/Artikelnr.",
    "Funktion", "Anmerkungen",
]

# Map Excel column index (0-based) to DB field name
COL_TO_FIELD = {
    0: "nr",
    1: "typ",
    2: "bezeichnung",
    3: "modell",
    4: "hersteller",
    5: "standort_name",
    6: "seriennummer",
    7: "mac_adresse",
    8: "ip_adresse",
    9: "firmware",
    10: "integration",
    11: "netzwerk",
    12: "stromversorgung",
    13: "ain_artikelnr",
    14: "funktion",
    15: "anmerkungen",
}


def _is_category_row(row_values: list[Any]) -> bool:
    """Detect category header rows: Nr cell is empty, merged across columns, bold text."""
    # Category rows have an empty first cell (Nr) and text in the second cell
    # with all remaining cells empty or same-merged
    if row_values[0] is not None and str(row_values[0]).strip():
        return False
    if not row_values[1] or not str(row_values[1]).strip():
        return False
    # Check that columns 3+ are empty (they are merged in the original)
    non_empty_after = sum(1 for v in row_values[3:] if v is not None and str(v).strip())
    return non_empty_after == 0


def _clean_value(val: Any) -> str | None:
    """Clean a cell value to string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse an Excel file in Geraeteuebersicht format and return device dicts.

    Returns a list of dicts ready for DB insertion (without id/uuid/timestamps).
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Find the right sheet
    ws = None
    for name in ["Geraeteuebersicht", "Sheet1", "Tabelle1"]:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    devices: list[dict[str, Any]] = []
    header_found = False
    current_category = ""

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        row_values = list(row)

        # Pad to 16 columns if needed
        while len(row_values) < 16:
            row_values.append(None)

        # Detect header row
        if not header_found:
            first_val = _clean_value(row_values[0])
            if first_val and first_val.lower() in ("nr", "nr."):
                header_found = True
                logger.info("Header found at row %d", row_idx)
            continue

        # Skip empty rows
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        # Detect category rows
        if _is_category_row(row_values):
            current_category = _clean_value(row_values[1]) or ""
            logger.debug("Category: %s (row %d)", current_category, row_idx)
            continue

        # Parse data row
        device: dict[str, Any] = {}
        for col_idx, field_name in COL_TO_FIELD.items():
            if col_idx < len(row_values):
                val = _clean_value(row_values[col_idx])
                if field_name == "nr" and val is not None:
                    try:
                        device[field_name] = int(float(val))
                    except (ValueError, TypeError):
                        device[field_name] = None
                else:
                    device[field_name] = val

        # Must have at least typ and bezeichnung
        if not device.get("typ") and not device.get("bezeichnung"):
            continue

        # Default typ if missing
        if not device.get("typ"):
            device["typ"] = "Sonstiges"
        if not device.get("bezeichnung"):
            device["bezeichnung"] = "Unbenannt"

        # Add uuid
        device["uuid"] = uuid4().hex

        devices.append(device)

    logger.info("Parsed %d devices from Excel", len(devices))
    return devices


def import_devices_to_db(conn, devices: list[dict[str, Any]]) -> int:
    """Insert parsed devices into the database. Returns count of inserted rows."""
    fields = [
        "uuid", "nr", "typ", "bezeichnung", "modell", "hersteller",
        "standort_name", "seriennummer", "mac_adresse", "ip_adresse",
        "firmware", "integration", "netzwerk", "stromversorgung",
        "ain_artikelnr", "funktion", "anmerkungen",
    ]
    placeholders = ", ".join(["?"] * len(fields))
    columns = ", ".join(fields)
    sql = f"INSERT INTO devices ({columns}) VALUES ({placeholders})"

    count = 0
    for device in devices:
        values = [device.get(f) for f in fields]
        try:
            conn.execute(sql, values)
            count += 1
        except Exception as e:
            logger.warning("Failed to insert device %s: %s", device.get("bezeichnung"), e)

    return count
