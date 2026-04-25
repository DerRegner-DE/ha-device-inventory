"""Excel export matching Geraeteuebersicht.xlsx format.

16 columns, dark blue header #1F4E79, alternating light blue #D6E4F0 / white rows,
category grouping, Arial 11, thin borders with light blue color.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- Style constants (matching create_device_list.py exactly) ---

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGN = Alignment(vertical="center", wrap_text=True)

ROW_FILL_ALT = PatternFill("solid", fgColor="D6E4F0")
ROW_FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")

CATEGORY_FONT = Font(name="Arial", bold=True, size=11, color="1F4E79")
CATEGORY_FILL = PatternFill("solid", fgColor="B4C6E7")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# v2.5.3: Field-to-label map + widths covers every exportable field. The
# frontend's ExportPicker lets the user pick any subset of these; previously
# the exporter silently ignored the ``fields=`` parameter because HEADERS
# was hardcoded to 16 columns and unused fields just came out as empty cells.
#
# Default column set (used when no ``fields`` is passed) matches the pre-v2.5.3
# behaviour so existing PDF/XLSX exports are visually unchanged.
FIELD_LABELS: dict[str, str] = {
    "nr": "Nr",
    "typ": "Typ",
    "bezeichnung": "Bezeichnung",
    "modell": "Modell",
    "hersteller": "Hersteller",
    "standort_name": "Standort",
    "standort_floor_id": "Etage",
    "standort_area_id": "Bereich-ID",
    "seriennummer": "Seriennummer",
    "ain_artikelnr": "AIN/Artikelnr.",
    "firmware": "Firmware",
    "integration": "Integration (HA)",
    "netzwerk": "Netzwerk",
    "stromversorgung": "Stromversorgung",
    "ip_adresse": "IP-Adresse",
    "mac_adresse": "MAC-Adresse",
    "anschaffungsdatum": "Anschaffung",
    "garantie_bis": "Garantie bis",
    "ha_device_id": "HA Device ID",
    "ha_entity_id": "HA Entity ID",
    "funktion": "Funktion",
    "anmerkungen": "Anmerkungen",
}

# Column widths in Excel units, keyed by DB field name.
FIELD_WIDTHS: dict[str, int] = {
    "nr": 5,
    "typ": 14,
    "bezeichnung": 32,
    "modell": 28,
    "hersteller": 20,
    "standort_name": 24,
    "standort_floor_id": 14,
    "standort_area_id": 22,
    "seriennummer": 22,
    "ain_artikelnr": 24,
    "firmware": 10,
    "integration": 22,
    "netzwerk": 14,
    "stromversorgung": 16,
    "ip_adresse": 14,
    "mac_adresse": 20,
    "anschaffungsdatum": 14,
    "garantie_bis": 14,
    "ha_device_id": 36,
    "ha_entity_id": 28,
    "funktion": 38,
    "anmerkungen": 36,
}

DEFAULT_FIELDS: list[str] = [
    "nr", "typ", "bezeichnung", "modell", "hersteller", "standort_name",
    "seriennummer", "mac_adresse", "ip_adresse", "firmware",
    "integration", "netzwerk", "stromversorgung", "ain_artikelnr",
    "funktion", "anmerkungen",
]

# Mapping from integration DB field to category label for grouping
INTEGRATION_CATEGORIES: list[tuple[str, list[str]]] = [
    ("FRITZ!Box Netzwerk", ["fritz", "fritzbox", "fritz, fritzbox"]),
    ("Zigbee (Zigbee2MQTT)", ["zigbee2mqtt", "zigbee2mqtt (MQTT)", "zha"]),
    ("Tuya (LocalTuya)", ["localtuya"]),
    ("Bosch Smart Home (SHC)", ["boschshc"]),
    ("HomeMatic IP", ["homematicip_cloud"]),
    ("Ring", ["ring"]),
    ("Blink", ["blink"]),
    ("Amazon Alexa Devices", ["alexa_devices", "alexa_media", "alexa_devices + alexa_media", "alexa_media (BT)"]),
    ("TP-Link", ["tplink"]),
    ("AVM Powerline", ["fritz (device_tracker)"]),
]


def _device_to_row(
    device: dict[str, Any], nr: int, fields: list[str]
) -> list[Any]:
    """Convert a device dict to a row with exactly the requested fields.

    ``nr`` is synthesised (not stored on the device), so the "nr" field is
    filled from the caller, not from the dict.
    """
    out: list[Any] = []
    for f in fields:
        if f == "nr":
            out.append(nr)
        else:
            out.append(device.get(f, ""))
    return out


def _categorize_devices(devices: list[dict[str, Any]]) -> list[tuple[str, list[dict[str, Any]]]]:
    """Group devices by integration category. Unmatched go to 'Sonstige Geraete'."""
    categorized: dict[str, list[dict[str, Any]]] = {}
    used_ids: set[int] = set()

    for cat_name, integrations in INTEGRATION_CATEGORIES:
        cat_devices = []
        for d in devices:
            integration = (d.get("integration") or "").strip().lower()
            if any(intg.lower() in integration or integration in intg.lower() for intg in integrations):
                cat_devices.append(d)
                used_ids.add(d["id"])
        if cat_devices:
            categorized[cat_name] = cat_devices

    # Remaining devices go to "Sonstige Geraete"
    remaining = [d for d in devices if d["id"] not in used_ids]
    if remaining:
        categorized["Sonstige Geraete"] = remaining

    # Return in defined order, then extras
    result = []
    for cat_name, _ in INTEGRATION_CATEGORIES:
        if cat_name in categorized:
            result.append((cat_name, categorized.pop(cat_name)))
    for cat_name, cat_devices in categorized.items():
        result.append((cat_name, cat_devices))

    return result


def export_devices_to_xlsx(
    devices: list[dict[str, Any]],
    fields: list[str] | None = None,
) -> bytes:
    """Export devices to Excel bytes.

    ``fields`` is the list of DB field names to include (in that order).
    When None, the default 16-column set from pre-v2.5.3 is used so existing
    user workflows don't suddenly change.
    """
    selected = [f for f in (fields or DEFAULT_FIELDS) if f in FIELD_LABELS]
    if not selected:
        selected = DEFAULT_FIELDS
    headers = [FIELD_LABELS[f] for f in selected]
    widths = [FIELD_WIDTHS.get(f, 18) for f in selected]

    wb = Workbook()
    ws = wb.active
    ws.title = "Geraeteuebersicht"

    # --- Header row ---
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # --- Group by category ---
    categories = _categorize_devices(devices)

    row = 2
    nr = 0

    for cat_name, cat_devices in categories:
        # Category header row
        for col in range(1, len(selected) + 1):
            c = ws.cell(row=row, column=col)
            c.fill = CATEGORY_FILL
            c.font = CATEGORY_FONT
            c.border = THIN_BORDER

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max(2, len(selected)))
        cell = ws.cell(row=row, column=2, value=cat_name)
        cell.font = CATEGORY_FONT
        cell.fill = CATEGORY_FILL
        row += 1

        # Data rows
        data_row_idx = 0
        for device in cat_devices:
            nr += 1
            values = _device_to_row(device, nr, selected)
            fill = ROW_FILL_ALT if data_row_idx % 2 == 0 else ROW_FILL_WHITE

            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = DATA_FONT
                c.alignment = DATA_ALIGN
                c.fill = fill
                c.border = THIN_BORDER

            row += 1
            data_row_idx += 1

    # --- Column widths ---
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Filters & freeze ---
    ws.auto_filter.ref = f"A1:{get_column_letter(len(selected))}{row - 1}"
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "1F4E79"

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
